#!/usr/bin/env python3
"""
Website Information Architecture Crawler

Crawls a website to map its navigation structure and generates reports.
Uses disk caching for fast incremental crawls.
"""

import argparse
import asyncio
import csv
import gzip
import hashlib
import re
import signal
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse
from urllib.robotparser import RobotFileParser

import aiohttp
import diskcache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from selectolax.parser import HTMLParser


# =============================================================================
# Configuration
# =============================================================================

USER_AGENT = "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)"
TIMEOUT = 8  # seconds - fail fast
CACHE_DAYS = 7

# CSS selectors for finding navigation elements
SELECTORS = {
    "nav": [
        "nav a",
        "[role='navigation'] a",
        "header a",
        ".nav a",
        ".menu a",
        ".navbar a",
        ".main-nav a",
    ],
    "sidebar": [
        ".sidebar a",
        "aside a",
        ".aside a",
        ".submenu a",
        ".section-nav a",
        ".qg-aside a",
    ],
    "breadcrumb": [
        ".breadcrumb a",
        ".breadcrumbs a",
        "[aria-label='breadcrumb'] a",
        ".qg-breadcrumb a",
    ],
    "footer": [
        "footer a",
        ".footer a",
        "#footer a",
    ],
}


# =============================================================================
# Data Classes
# =============================================================================


class Page:
    """Information about a crawled page."""

    def __init__(self, url, title="", depth=0, parent="", from_cache=False):
        self.url = url
        self.title = title
        self.depth = depth  # crawl depth
        self.parent = parent
        self.from_cache = from_cache
        self.in_nav = False
        self.in_sitemap = False
        self.inbound_links = 0
        self.breadcrumb = []

    @property
    def url_depth(self):
        """Depth based on URL path segments."""
        path = urlparse(self.url).path.strip("/")
        return len(path.split("/")) if path else 0


# =============================================================================
# Cache
# =============================================================================

_cache = None


def init_cache(cache_dir):
    """Initialize disk cache."""
    global _cache
    _cache = diskcache.Cache(str(cache_dir), size_limit=500 * 1024 * 1024)


def cache_get(url):
    """Get cached HTML."""
    if _cache:
        return _cache.get(hashlib.md5(url.encode()).hexdigest())
    return None


def cache_set(url, html):
    """Cache HTML."""
    if _cache:
        _cache.set(
            hashlib.md5(url.encode()).hexdigest(), html, expire=CACHE_DAYS * 86400
        )


# =============================================================================
# HTTP Fetching
# =============================================================================


async def fetch(session, url):
    """
    Fetch URL content, using cache if available.
    Returns (html, from_cache) or (None, False) on error.
    """
    # Try cache first
    cached = cache_get(url)
    if cached:
        return cached, True

    # Fetch from network
    try:
        async with session.get(
            url, timeout=aiohttp.ClientTimeout(total=TIMEOUT)
        ) as resp:
            if resp.status == 200:
                html = await resp.text()
                cache_set(url, html)
                return html, False
    except Exception:
        pass
    return None, False


async def fetch_robots(session, base_url):
    """Fetch and parse robots.txt."""
    rp = RobotFileParser()
    rp.set_url(urljoin(base_url, "/robots.txt"))

    html, _ = await fetch(session, urljoin(base_url, "/robots.txt"))
    if html:
        rp.parse(html.splitlines())
    else:
        # No robots.txt = allow everything
        rp.parse(["User-agent: *", "Allow: /"])

    return rp


async def fetch_sitemap(session, url, visited=None):
    """Fetch sitemap URLs recursively."""
    if visited is None:
        visited = set()
    if url in visited:
        return []
    visited.add(url)

    try:
        async with session.get(
            url, timeout=aiohttp.ClientTimeout(total=TIMEOUT)
        ) as resp:
            if resp.status != 200:
                return []
            content = await resp.read()
            if url.endswith(".gz"):
                content = gzip.decompress(content)
            text = content.decode("utf-8")
    except Exception:
        return []

    urls = []
    if "<sitemapindex" in text.lower():
        # Sitemap index - recurse into nested sitemaps
        for match in re.finditer(r"<loc>\s*(.*?)\s*</loc>", text, re.I):
            urls.extend(await fetch_sitemap(session, match.group(1).strip(), visited))
    else:
        # Regular sitemap
        urls = [
            m.group(1).strip()
            for m in re.finditer(r"<loc>\s*(.*?)\s*</loc>", text, re.I)
        ]

    return urls


# =============================================================================
# HTML Parsing
# =============================================================================


def extract_links(html, base_url, selectors):
    """
    Extract links from HTML using CSS selectors.
    Returns list of (url, text) tuples.
    """
    parser = HTMLParser(html)
    base_host = urlparse(base_url).netloc
    links = []
    seen = set()

    for selector in selectors:
        for node in parser.css(selector):
            href = node.attributes.get("href", "")

            # Skip empty, anchor, and javascript links
            if not href or href.startswith("#") or href.startswith("javascript:"):
                continue

            # Build full URL
            full_url = urljoin(base_url, href)
            parsed = urlparse(full_url)

            # Skip non-http and external links
            if parsed.scheme not in ("http", "https"):
                continue
            if parsed.netloc != base_host:
                continue

            # Normalize URL (strip fragment)
            clean_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
            if parsed.query:
                clean_url += f"?{parsed.query}"

            if clean_url not in seen:
                seen.add(clean_url)
                text = (node.text(strip=True) or "")[:100]
                links.append((clean_url, text))

    return links


def extract_title(html):
    """Extract page title."""
    parser = HTMLParser(html)
    for selector in ["title", "h1"]:
        node = parser.css_first(selector)
        if node:
            return node.text(strip=True)[:200]
    return ""


def extract_all_links(html, base_url):
    """Extract ALL internal links (for inbound counting)."""
    return set(url for url, _ in extract_links(html, base_url, ["a"]))


# =============================================================================
# Main Crawler
# =============================================================================

_shutdown = False


def handle_signal(sig, frame):
    global _shutdown
    print("\n[!] Stopping...")
    _shutdown = True


async def crawl(start_url, max_depth=3, max_pages=500, concurrency=10, cache_dir=None):
    """
    Crawl a website and return structured data.

    Args:
        start_url: URL to start crawling from
        max_depth: Maximum link depth to follow
        max_pages: Maximum pages to crawl
        concurrency: Number of concurrent requests
        cache_dir: Directory for disk cache (None to disable)

    Returns:
        Dict with pages, nav_links, stats, etc.
    """
    global _shutdown
    _shutdown = False

    # Setup
    base_url = f"{urlparse(start_url).scheme}://{urlparse(start_url).netloc}"
    if cache_dir:
        init_cache(cache_dir)
        print(f"  [+] Cache: {cache_dir}")

    print(f"\n[*] Crawling {base_url}")
    print(f"    Depth: {max_depth}, Max pages: {max_pages}, Concurrent: {concurrency}")

    # HTTP session
    connector = aiohttp.TCPConnector(limit=concurrency)
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }

    async with aiohttp.ClientSession(connector=connector, headers=headers) as session:
        # Fetch robots.txt
        robots = await fetch_robots(session, base_url)
        print(f"  [+] Loaded robots.txt")

        # Try to find sitemap
        sitemap_urls = []
        for sitemap_url in [f"{base_url}/sitemap.xml", f"{base_url}/sitemap_index.xml"]:
            sitemap_urls = await fetch_sitemap(session, sitemap_url)
            if sitemap_urls:
                print(f"  [+] Found {len(sitemap_urls)} URLs in sitemap")
                break
        sitemap_set = set(sitemap_urls)

        # Data structures
        pages = {}
        nav_links = []  # (url, text, section, source_page)
        inbound_counts = defaultdict(int)
        link_sources = defaultdict(set)

        # BFS queue: (url, depth, parent_url)
        queue = [(start_url, 0, "")]
        visited = set()
        crawled = 0
        cached_hits = 0

        # Crawl loop
        while queue and crawled < max_pages and not _shutdown:
            # Build batch of URLs to fetch
            batch = []
            while queue and len(batch) < concurrency:
                url, depth, parent = queue.pop(0)

                if url in visited or depth > max_depth:
                    continue
                if not robots.can_fetch(USER_AGENT, url):
                    continue

                visited.add(url)
                batch.append((url, depth, parent))

            if not batch:
                break

            # Fetch batch concurrently
            tasks = [fetch(session, url) for url, _, _ in batch]
            results = await asyncio.gather(*tasks)

            # Process results
            for (url, depth, parent), (html, from_cache) in zip(batch, results):
                if not html:
                    continue

                # Create page
                page = Page(url, extract_title(html), depth, parent, from_cache)
                page.in_sitemap = url in sitemap_set
                pages[url] = page

                crawled += 1
                if from_cache:
                    cached_hits += 1

                # Progress
                indicator = "C" if from_cache else "N"
                print(f"  [{depth}]{indicator} {url[:70]}")

                # Extract navigation links
                for section, selectors in SELECTORS.items():
                    # Only extract footer from homepage
                    if section == "footer" and depth > 0:
                        continue
                    for link_url, link_text in extract_links(html, base_url, selectors):
                        nav_links.append((link_url, link_text, section, url))

                # Track inbound links
                for link_url in extract_all_links(html, base_url):
                    inbound_counts[link_url] += 1
                    link_sources[link_url].add(url)

                # Queue new URLs
                if depth < max_depth:
                    crawl_selectors = SELECTORS["nav"] + SELECTORS["sidebar"]
                    for link_url, _ in extract_links(html, base_url, crawl_selectors):
                        if link_url not in visited:
                            queue.append((link_url, depth + 1, url))

            # Small delay between batches
            await asyncio.sleep(0.05)

        # Update pages with inbound counts
        for url, count in inbound_counts.items():
            if url in pages:
                pages[url].inbound_links = count

        # Mark pages that appear in navigation
        for url, _, _, _ in nav_links:
            if url in pages:
                pages[url].in_nav = True

        print(f"\n[+] Done! {crawled} pages ({cached_hits} from cache)")

        return {
            "pages": pages,
            "nav_links": nav_links,
            "sitemap_urls": sitemap_urls,
            "base_url": base_url,
            "stats": {
                "crawled": crawled,
                "cached": cached_hits,
                "nav_links": len(nav_links),
            },
            "inbound_counts": dict(inbound_counts),
            "link_sources": {k: list(v) for k, v in link_sources.items()},
        }


# =============================================================================
# Report Generation
# =============================================================================


def build_tree(pages):
    """Build hierarchical tree from URL paths."""
    tree = {"_children": {}, "_pages": []}

    for url, page in pages.items():
        path = urlparse(url).path.strip("/")
        segments = path.split("/") if path else []

        node = tree
        for seg in segments:
            if seg not in node["_children"]:
                node["_children"][seg] = {"_children": {}, "_pages": []}
            node = node["_children"][seg]

        title = page.title.split("|")[0].strip()[:60] if page.title else ""
        node["_pages"].append({"url": url, "title": title})

    return tree


def render_tree(tree, indent=0, max_depth=3):
    """Render tree as markdown list."""
    lines = []
    prefix = "  " * indent

    for info in tree["_pages"]:
        title = info["title"] or urlparse(info["url"]).path or "/"
        lines.append(f"{prefix}- **{title}**")

    if indent < max_depth:
        for seg, subtree in sorted(tree["_children"].items()):
            # Use first page title as section title
            seg_title = seg
            if subtree["_pages"]:
                seg_title = subtree["_pages"][0]["title"] or seg
            lines.append(f"{prefix}- **{seg_title}**")
            lines.extend(render_tree(subtree, indent + 1, max_depth))

    return lines


def generate_markdown(data, path):
    """Generate markdown report."""
    domain = urlparse(data["base_url"]).netloc
    stats = data["stats"]

    lines = [
        f"# Site IA: {domain}",
        "",
        f"Crawled: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## Statistics",
        "",
        "| Metric | Count |",
        "|--------|-------|",
        f"| Pages crawled | {stats['crawled']} |",
        f"| From cache | {stats['cached']} |",
        f"| Nav links found | {stats['nav_links']} |",
        "",
        "## Most Linked Pages",
        "",
        "| Links | Path | Title |",
        "|-------|------|-------|",
    ]

    # Top 20 most linked pages
    top = sorted(data["inbound_counts"].items(), key=lambda x: -x[1])[:20]
    for url, count in top:
        page = data["pages"].get(url)
        title = page.title.split("|")[0][:35] if page and page.title else "-"
        path_str = urlparse(url).path or "/"
        lines.append(f"| {count} | `{path_str[:40]}` | {title} |")

    lines.append("")
    lines.append("## Information Architecture")
    lines.append("")

    # Build and render tree
    tree = build_tree(data["pages"])
    for seg, subtree in sorted(tree["_children"].items()):
        seg_title = subtree["_pages"][0]["title"] if subtree["_pages"] else seg
        lines.append(f"### {seg_title}")
        lines.append("")
        tree_lines = render_tree(subtree, 0, 3)[:40]
        lines.extend(tree_lines)
        lines.append("")

    # Navigation by section
    lines.append("## Navigation Links")
    lines.append("")

    by_section = defaultdict(set)
    for url, text, section, _ in data["nav_links"]:
        by_section[section].add((url, text))

    for section, links in by_section.items():
        lines.append(f"### {section.title()} ({len(links)} unique)")
        lines.append("")
        for url, text in list(links)[:25]:
            lines.append(f"- [{text or url}]({url})")
        if len(links) > 25:
            lines.append(f"- ... {len(links) - 25} more")
        lines.append("")

    path.write_text("\n".join(lines))
    print(f"  [+] {path}")


def generate_csv(data, path):
    """Generate CSV report."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["url", "title", "depth", "inbound_links", "in_nav", "in_sitemap"]
        )

        for page in sorted(
            data["pages"].values(), key=lambda p: (-p.inbound_links, p.url)
        ):
            writer.writerow(
                [
                    page.url,
                    page.title,
                    page.url_depth,
                    page.inbound_links,
                    page.in_nav,
                    page.in_sitemap,
                ]
            )

    print(f"  [+] {path}")


def generate_xlsx(data, path):
    """Generate Excel report."""
    wb = Workbook()
    header_style = {
        "fill": PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        ),
        "font": Font(color="FFFFFF", bold=True),
    }

    # URLs sheet
    ws = wb.active
    assert ws is not None
    ws.title = "URLs"
    ws.append(["URL", "Title", "Depth", "Inbound Links", "In Nav"])
    for cell in ws[1]:
        cell.fill = header_style["fill"]
        cell.font = header_style["font"]

    for page in sorted(data["pages"].values(), key=lambda p: (-p.inbound_links, p.url)):
        ws.append(
            [
                page.url,
                page.title[:80],
                page.url_depth,
                page.inbound_links,
                "Y" if page.in_nav else "",
            ]
        )

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 40

    # Hub pages sheet
    ws2 = wb.create_sheet("Hub Pages")
    ws2.append(["URL", "Inbound Links", "Title"])
    for cell in ws2[1]:
        cell.fill = header_style["fill"]
        cell.font = header_style["font"]

    for url, count in sorted(data["inbound_counts"].items(), key=lambda x: -x[1])[:100]:
        page = data["pages"].get(url)
        ws2.append([url, count, page.title[:60] if page else ""])

    ws2.column_dimensions["A"].width = 60

    wb.save(path)
    print(f"  [+] {path}")


def generate_reports(data, output_dir):
    """Generate all reports."""
    output_dir.mkdir(parents=True, exist_ok=True)
    domain = urlparse(data["base_url"]).netloc.replace(".", "-")

    print("\n[*] Generating reports...")
    generate_markdown(data, output_dir / f"{domain}.md")
    generate_csv(data, output_dir / f"{domain}.csv")
    generate_xlsx(data, output_dir / f"{domain}.xlsx")


# =============================================================================
# CLI
# =============================================================================


def main():
    signal.signal(signal.SIGINT, handle_signal)
    signal.signal(signal.SIGTERM, handle_signal)

    parser = argparse.ArgumentParser(
        description="Crawl a website and map its information architecture"
    )
    parser.add_argument("url", help="URL to crawl")
    parser.add_argument(
        "-d", "--depth", type=int, default=3, help="Max depth (default: 3)"
    )
    parser.add_argument(
        "-m", "--max-pages", type=int, default=500, help="Max pages (default: 500)"
    )
    parser.add_argument(
        "-c",
        "--concurrency",
        type=int,
        default=10,
        help="Concurrent requests (default: 10)",
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=Path(__file__).parent / "reports"
    )
    parser.add_argument("--no-cache", action="store_true", help="Disable disk cache")

    args = parser.parse_args()

    url = args.url if args.url.startswith("http") else f"https://{args.url}"
    cache_dir = None if args.no_cache else Path(__file__).parent / ".cache"

    data = asyncio.run(
        crawl(url, args.depth, args.max_pages, args.concurrency, cache_dir)
    )
    generate_reports(data, args.output)

    print(f"\n[+] Reports saved to {args.output}/")


if __name__ == "__main__":
    main()
